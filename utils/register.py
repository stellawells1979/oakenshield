

'''
管理签到消息
'''
import json
import copy
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta
import config
from utils import timestand
from utils.tools import tools
from database import sql
from logmanage import DailyLogManager
from utils.quick import quick
log = DailyLogManager('Register', logging.ERROR, logging.INFO)

class Register:
    '''
    签到模块
    '''
    def __init__(self):

        self.table = sql.table_register
        self.support_param = {
            # 当机器人识别到用户需要输入参数时向用户提示输入参数
            'describe': '请输入签到描述，长度100字以内',
            'period': '请输入签到周期，以天为单位，最多设置31天的签到周期',
            'regi_name': '你正在创建新的签到项目，请输入签到项目名称，长度20字以内',
        }
        self.send_file = None

    def register_status(self, group, rules):
        '''
        维护签到数据表的签到状态，确保签到状态保持在有效的时间内
        并将此状态同步到规则参数
        :param group:
        :param rules: 签到规则参数，用于更新签到状态
        :return: 返回更新后的签到规则参数
        '''

        # 从数据表中获取指定群组的所有签到项目
        query = f"SELECT `id`,`expired`,`status` FROM `{self.table}` WHERE `chat`=%s"
        all_register = sql.query(sql.database, query, [group]) or []

        for row in all_register:
            if not row:
                continue

            # 签到状态的三种，Run:进行中，Exp:已过期，End:强制关闭，
            status = row.get('status')
            if status != 'End' and timestand.date_from_timestamp() > row.get('expired'):
                # 将签到数据表中的当前签到项目更新为过期状态
                query = f"UPDATE `{sql.table_register}` SET `status`=%s WHERE `id`=%s"
                sql.query(sql.database, query, ['Exp', row.get('id')])
                if row.get('id') == rules.get('regi_id'):
                    # 重置签到规则为默认状态
                    rules = copy.deepcopy(config.rules_example['register'])

        rules.update({'regi_count': len(all_register)})

        return rules

    def create_register(self, group, user, rules):
        '''
        在数据库中创建一个签到实例
        '''
        regi_id = None
        while not regi_id:
            # 为当前签到项目创建一个ID并确保不会生成重复的短ID
            regi_id = tools.brief_uid(6)
            query = f"SELECT `id` FROM `{self.table}` WHERE `id`=%s"
            if sql.query(sql.database, query, [regi_id]):
                regi_id = None

        params = {
            'id': regi_id,
            'chat': group,
            'creator': user,
            'name': rules['regi_name'],
            'describe': rules['describe'],
            'origin': rules['origin'],
            'expired': rules['expired'],
            'timing': rules['timing'],
            'status': 'Run'
        }
        fields = [f"`{row}`" for row in params.keys()]
        query = f'INSERT INTO `{self.table}` ({",".join(fields)}) VALUES ({",".join(["%s"] * len(params))})'
        sql.query(sql.database, query, list(params.values()))

        rules.update({'regi_count': rules.get('regi_count') + 1})
        rules.update({'regi_id': regi_id})
        rules.update({'status': 'Run'})
        if not rules.get('timing'):
            rules.update({'timing': 30})

        return rules

    def parse_register(self, regi_id, group_id, users, reg_date, message_id):
        '''
        解析签到消息
        :param regi_id: 签到ID
        :param group_id:
        :param users:
        :param reg_date:
        :param message_id:
        :return:
        '''

        query = f"SELECT * FROM `{self.table}` WHERE `chat`=%s and `id`=%s"
        chat_register = sql.query(sql.database, query, [group_id, regi_id]) or []


        if not chat_register:
            return []

        # 检查当前签到项目是否进行中
        chat_register = chat_register[0]
        if chat_register.get('status') != 'Run':
            return []

        current_day = timestand.date_from_timestamp()     # 获取当前日期(基于项目指定的时区)
        register_day = timestand.date_from_timestamp(reg_date)   # 获取签到日期，即消息日期(基于项目指定的时区)

        if current_day != register_day:

            log.warning(f"签到日期:{register_day}，当前日期{current_day}, 这是作弊")
            return []

        expired = chat_register.get('expired')  # 从签到数据提取到期日期

        if register_day > expired:
            query = f"UPDATE {self.table} SET `expired`=%s WHERE `id`=%s"
            sql.query(sql.database, query, ['Exp', regi_id])

            return []


        result = []

        # 获取签到起始日期并计算出当前签到日，提取当日签到数据
        origin = register_day - chat_register.get('origin') # 获取签到起始日期
        register_field = f'RE_{str(origin.days + 1)}'     # 匹配当前签到日期字段
        current_register = chat_register.get(register_field) or {}   # 提取当日签到数据

        # 提取当前用户签到数据
        user_id = users.get('id')
        user_register = current_register.get(str(user_id)) or {}

        if not user_register:
            cumulative = 0
            for row in range(1, origin.days + 1):
                if chat_register.get(f'RE_{row}') and chat_register.get(f'RE_{row}').get(str(user_id)):
                    cumulative = cumulative + 1
            if cumulative == origin.days:
                send_text = f"🎉🎉 签到成功\n🎉🎉你已连续签到{cumulative + 1}天\n\n{chat_register.get('describe')}"
            else:
                send_text = f"🎉🎉 签到成功\n🎉🎉你已累计签到{cumulative + 1}天\n\n{chat_register.get('describe')}"

            current_register.update({
                str(user_id):{
                    **({'username': users.get('username')} if users.get('username') else {}),
                    **({'first_name': users.get('first_name')} if users.get('first_name') else {}),
                    **({'last_name': users.get('last_name')} if users.get('last_name') else {}),
                    'register': [str(timestand.format_datetime_tz(reg_date))]
                }
            })

        else:

            register_count = user_register.get('register')
            if len(register_count) > 2:
                mute_time = reg_date + 600
                send_text = '🈲 因你涉嫌恶意签到，已被禁言10分钟'
                if len(register_count) > 4:
                    mute_time = reg_date + (24 * 600)
                    send_text = '🈲🈲 因你涉嫌恶意签到且超过五次，已被禁言24小时'

                result.append([
                    'restrictChatMember',
                    {'chat_id': group_id, 'user_id': user_id, 'permissions': quick.permissions(True), 'until_date': mute_time},
                    None
                ])
            else:
                send_text = '⚠️⚠️ 请勿重复签到'
            register_count.append(str(timestand.format_datetime_tz(reg_date)))

        query = f"UPDATE `{self.table}` SET `{register_field}`=%s WHERE `id`=%s"
        sql.query(sql.database, query, [json.dumps(current_register, ensure_ascii=False), regi_id])

        result.append([
            'sendMessage',
            {'chat_id': group_id, 'text': send_text, 'reply_to_message_id': message_id},
            None
        ])

        return result

    def set_register(self, user, group, message_id, rules, details):
        '''
        设置签到规则
        :param user: 当前用户ID，如果用户启动了签到项目，会调用 create_register 方法创建签到数据表，此时会用到这个ID参数
        :param group: 当前群组ID，当用户在设置自定义参数时，需要构建 waitinput 并记录到数据库
        :param message_id: 消息ID，当用户在设置自定义参数时，需要构建 waitinput 并记录到数据库
        :param rules: 当前规则
        :param details: 用户正在设置的规则参数
        :return:
        '''
        extra_text = None
        waitinput = None
        if details == 'regi_name':

            if rules.get('status') == 'Run':
                extra_text = '⚠️ 一个群组只能运行一个活动的签到项目'
            else:
                # 创建新签到项目时，将其它参数置空
                for key, value in rules.items():
                    if key in ['regi_count']:
                        continue
                    rules.update({key: None})
                extra_text = self.support_param.get(details)
                waitinput = f'rules|register|{details}|{group}|{message_id}'

        elif details in ['describe', 'period'] and rules.get('status') in ['Run', 'End', 'Exp']:
            # 如果用户尝试更改签到规则，则检查当前签到规则是否正在进行，如果是，则提示用户无法更改进行中的签到规则
            extra_text = {
                'Run': '⚠️ 无法修改正在进行中的签到项目',
                'End': '⚠️ 当前签到项目已关闭，你可以创建新的签到项目',
                'Exp': '⚠️ 当前签到项目已过期，你可以创建新的签到项目',
            }.get(rules.get('status'))

        elif details in ['describe', 'period']:

            if not rules.get('regi_name'):
                # 只有创建了项目名称才允许废黜其它参数
                extra_text = '⚠️ 请先创建签到项目'
            else:
                waitinput = f'rules|register|{details}|{group}|{message_id}'
                extra_text = self.support_param.get(details)

        elif details == 'timing':
            extra_text = '⚠️ 请输入签到轮循时间，以分钟为单位，最少10分钟，最多120分钟，默认30分钟'
            waitinput = f'rules|register|{details}|{group}|{message_id}'

        elif details == 'begin':
            rules = self.create_register(group, user, rules)
            extra_text = '签到项目已启动'


        elif details == 'End':
            rules.update({'status': 'End'})
            query = f"UPDATE `{sql.table_register}` SET `status`=%s WHERE `id`=%s"
            sql.query(sql.database, query, ['End', rules.get('regi_id')])
            extra_text = '已关闭当前签到项目'

        if waitinput:
            query = (f'INSERT INTO {sql.table_interact} (bot,user,waitinput) VALUES (%s,%s,%s) '
                     f'ON DUPLICATE KEY UPDATE waitinput=%s, edited=NOW()')
            sql.query(sql.database, query, ['rules', user, waitinput, waitinput])

        return rules, extra_text

    def set_regiser_param(self, rules, details, param):
        '''
        设置签到规则的自定义参数

        :param rules: 签到规则
        :param details: 参数名称
        :param param: 参数
        :return:
        '''

        if details == 'period':
            rules.update({'origin': str(timestand.today())})
            rules.update({'expired': str(timestand.today() + timedelta(days=param))})
        else:
            rules.update({details: param})

        if rules['status'] == 'Run' and details == 'timing':
            query = f"UPDATE {self.table} SET `timing`=%s WHERE `id`=%s"
            sql.query(sql.database, query, [param, rules.get('regi_id')])

        elif rules['status'] is None and rules['regi_name'] and rules['describe'] and rules['origin'] and rules['expired']:
            rules.update({'status': 'Read'})

        return rules

    def history(self, chat, group, regi_id):
        '''
        响应用户对历史签到项目的操作
        :param chat:
        :param group:
        :param regi_id: 签到项目ID(6位)，如果此参数以 ? 或 ! 结尾，则是要求提取该项目的全部详情或者是清除该签到项目的操作
        :return:
        '''
        regi_id = regi_id if regi_id != '0' else None
        extra_text = '点击相应项目可查看签到详情'    # 初始化说明文本

        extra = None    # 初始化额外操作,此参数在regi_id中提取，通常情况下regi_id由5个字符构成，如果regi_id是8个字符，则第6个字符就是额外操作参数
        if regi_id and len(regi_id) > 6:
            # 检查regi_id长度提取可能存在的额外操作，通常是!或?,分别表示查看当前签到ID的详情和删除当前签到ID
            regi_id, extra = regi_id[:6], regi_id[6:]

        if extra == '!':
            # 删除当前签到数据, 首先查询当前签到项目状态是否正在真行中
            query = f"SELECT `name`,`status` FROM `{self.table}` WHERE `id`=%s"
            curr_status = sql.query(sql.database, query, [regi_id]) or []
            if curr_status and curr_status[0]['status'] == 'Run':
                extra_text = '无法删除正在进行中的签到项目，请先停止该项目再重试'
            else:
                # 从数据库删除当前签到项目
                extra_text = f"已清除签到项目：【{curr_status[0].get('name')}】"
                # query = f"DELETE FROM `{self.table}` WHERE `id`=%s"
                # sql.query(sql.database, query, [regi_id])

        elif regi_id:
            # 获取指定签到项目详情，并返回相应参数
            return self.history_node(chat, group, regi_id, extra)

        # 列出当前群组的所有签到项目
        query = f"SELECT `id`,`name` FROM `{self.table}` WHERE `chat`=%s"
        historys = sql.query(sql.database, query, [group]) or []

        buttons = []
        for row in historys:
            if not row:
                continue
            buttons.append({'text': row['name'], 'callback_data': f"rules|register|history|{row.get('id')}|{group}"})
        guid_button = [{'text': '返回', 'callback_data': f'rules|register|0|1|{group}'}]
        param_text = f"当前群组共有{len(historys)}个签到项目" if len(historys) > 0 else '当前群组没有签到项目'

        # 返回主键盘，导航键盘，参数文本，额外文本，文件对象（点位符）
        return buttons, guid_button, param_text, extra_text, None

    def history_node(self, chat, group, regi_id, extra):
        '''
        :param chat: 群聊id
        :param group:
        :param regi_id:
        :param extra:
        :return:
        '''
        file_obj = None     # 初始化文件对象

        # 从数据库中查询签到项目信息
        query = f"SELECT * FROM `{self.table}` WHERE `id`=%s"
        result = sql.query(sql.database, query, [regi_id]) or []
        if not result:
            return []

        # 提取签到项目各参数
        result = result[0]
        origin = result.get('origin')      # 起始日期
        expired = result.get('expired')        # 结束日期
        days_count = (expired - origin).days    # 持续天数

        # 拆分历史数据，提取签到项目信息和签到周期内每一天的签到数据
        register_info = {}
        history_data = {}
        register_dates = []
        for row in list(result.keys()):
            if row.startswith('RE_') and days_count >= int(row[3:]):
                register_date = (origin + timedelta(days=(int(row[3:]) - 1))).strftime('%Y-%m-%d')
                history_data.update({register_date: result.pop(row)})
                register_dates.append(register_date)
            elif not row.startswith('RE_'):
                register_info.update({row: result.pop(row)})

        # 重构签到数据,原始的数据表是以签到日期为健，重构后将是以用户为健的签到数据（dict)
        history_data = self.reorganize_register(history_data)

        # 迭代签到数据，统计满签用户，
        continuous = 0
        for key, value in history_data.items():
            if value and len(value) > days_count:
                # value包含了用户信息及签到记录，days_count是签到项目的总天数
                continuous = continuous + 1

        # 构建签到项目的概述文本，此文本将以rules中的参数文本格式向用户展示
        param_text = (f"项目名称：{register_info.get('name')}\n签到说明：{register_info.get('describe')}\n起始日期：{origin}\n"
                      f"结束日期：{expired}\n连续满签人数：{continuous} 人") \

        # 设置额外的说明文本
        extra_text = '点击【获取详细签到数据】，机器人将发送一分Excel文档'

        if extra == '?':
            # 响应生成详细签到数据的请求，并重置额外说明文本
            extra_text = '机器人已向你发送了一个Excel文件，请注意查收'
            file_obj = self.create_excel(register_info, register_dates, history_data, chat)

        # 创建功能按钮，此按钮将以rules中的主按钮格式向用户展示
        buttons = [
            {'text': '获取详细签到数据', 'callback_data': f"rules|register|history|{regi_id}?|{group}"},
            {'text': '清除当前签到数据', 'callback_data': f"rules|register|history|{regi_id}!|{group}"},
        ]
        guid_button = [{'text': '返回', 'callback_data': f'rules|register|history|0|{group}'}]

        # 返回主键盘，导航键盘，参数文本，额外文本，文件对象
        return buttons, guid_button, param_text, extra_text, file_obj

    @classmethod
    def create_excel(cls, info, dates, data, chat):
        '''
        创建csv文档

        :param info: 签到项目的基本信息
        :param dates: 历史签到日期
        :param data: 历史签到数据
        :param chat: 群聊id
        :return:
        '''
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "数据表"

        # 根据 headers 动态设置列宽
        headers = ['id', 'username', 'first_name', 'last_name', *dates]
        for col_index, header in enumerate(headers, start=1):
            column_letter = get_column_letter(col_index)

            if header in ['id', 'username', 'first_name', 'last_name']:
                width = 18
            else:
                width = 28
            ws.column_dimensions[column_letter].width = width

        # 设置第一，第二行的行高
        ws.row_dimensions[1].height = 150
        ws.row_dimensions[2].height = 24

        # 统筹字体样式
        header_font = Font(name="微软雅黑", size=12, bold=True)
        body_font = Font(name="微软雅黑", size=10)
        title_font = Font(name="微软雅黑", size=10, bold=True)

        # 统筹对齐，换行样式 horizontal：相对于单元格的垂直对齐，vertical：相对于单元格的水平对齐，wrap_text 表示是否换行
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)      # 水平垂直居中样式
        left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)   # 水平居中，垂直居左样式
        # left_top = Alignment(horizontal="left", vertical="top", wrap_text=True) # 水平居上，垂直居左样式

        thin_side = Side(style="thin", color="000000")
        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        # 构建概述内容，合并第1行 B 到最后一列作为单元格以填充签到概述内容
        title_text = ''
        for key, value in info.items():
            if key in ['chat', 'scheme', 'edited', 'created']:
                continue
            title_text = title_text + f"{config.translation.get(key, key)}: {value}\n"

        # 构建第一行第一列样式并填充文本
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value='签到概述')
        title_cell.font = title_font
        title_cell.alignment = left_center
        title_cell.border = thin_border

        # 构建第一行第二列样式并填充文本
        title_content_cell = ws.cell(row=1, column=2, value=title_text.rstrip('\n'))
        title_content_cell.font = title_font
        title_content_cell.alignment = left_center
        title_content_cell.border = thin_border

        # 合并单元格的边框不会自动完整应用到整个区域， 因此需要给第一行所有涉及的单元格都设置边框。
        for col_index in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_index).border = thin_border


        # 定义表头（第2行）
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_index, value=header)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

        # 构建表头到列号的映射，方便根据字段名或日期定位列
        header_index = {header: index for index, header in enumerate(headers, start=1)}

        # 迭代历史签到数据，将签到信息统计到excel数据表
        # {1120690440: [{'username': 'bigapple699','last_name': '苹果','first_name': '大'},['2026-06-09 15:41:51 UTC+8','2026-06-09 21:42:30 UTC+8'],...]}
        # 上述示例数据对象，索引1是固定的用户信息，从索引2开始是签到日期数据，每个索引对应当日签到记录，当日可能会包含多个签到记录
        for row_index, (user_id, row_data) in enumerate(data.items(), start=3):
            user_info = row_data[0] if row_data and isinstance(row_data[0], dict) else {}

            # 初始化一行数据，先按表头全部初始化为空
            excel_row = {header: '' for header in headers}

            # 填充用户基础信息
            excel_row['id'] = user_id
            excel_row['username'] = user_info.get('username', '')
            excel_row['first_name'] = user_info.get('first_name', '')
            excel_row['last_name'] = user_info.get('last_name', '')

            # 填充签到日期数据
            for register_list in row_data[1:]:
                if not register_list:
                    continue

                # register_list 示例：
                # ['2026-06-09 15:41:51 UTC+8', '2026-06-09 21:42:30 UTC+8']
                if isinstance(register_list, list):
                    register_date = register_list[0][:10]

                    if register_date in excel_row:
                        excel_row[register_date] = '\n'.join(register_list)

            # 根据当前行中实际文本的最大行数设置行高
            max_line_count = 1
            for value in excel_row.values():
                if isinstance(value, str) and value:
                    line_count = value.count('\n') + 1
                    max_line_count = max(max_line_count, line_count)

            ws.row_dimensions[row_index].height = max(24, max_line_count * 18)

            # 按 headers 顺序写入 Excel
            for header in headers:
                col_index = header_index[header]
                cell = ws.cell(row=row_index, column=col_index, value=excel_row[header])
                cell.font = body_font
                cell.alignment = center
                cell.border = thin_border

        wb.save(f"test/{info.get('id')}_register_history.xlsx")

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)


        # result = [
        #     'sendDocument',
        #     {'chat_id': chat},
        #     {
        #         'file': {
        #             'type': 'document',
        #             'path': f"test/{info.get('id')}_register_history.xlsx",
        #             'filename': f"{info.get('id')}_register_history.xlsx"
        #         }
        #     }
        # ]

        # 直接返回一个telegram bot 请求对象
        return [
            'sendDocument',
            {'chat_id': chat},
            {
                'file': {
                    'type': 'document',
                    'file': excel_file,
                    'filename': f"{info.get('id')}_register_history.xlsx"
                }
            }
        ]

    @classmethod
    def reorganize_register(cls, data):
        '''
        重构签到数据,原始的数据表是以签到日期为健，重构后将返回以用户为健的数据（dict)
        :return:
        '''
        # 首先获取所有的签到用户，并初始化化一个以用户为键，签到记录为空列表的数据字典，示例： {user1: [], user2: [].....}
        result = {}
        for key, value in data.items():
            if not value:
                continue
            for user, regi in value.items():
                result.update({user: []})
        # result 的数据结构：{user1: [], user2: [].....}

        for user, regi in result.items():
            users = None    # 初始化用户信息
            for key, value in data.items():
                # 迭代签到数据，提取当前用户的签到记录，添加到用户数据中
                if not value:
                    continue
                if user in value:
                    regi.append(value.get(user).pop('register'))
                    users = value.get(user)
            if users:
                # 如果用户信息有效，则插入到用户数据的列首
                regi.insert(0, users)

        return result

def create_register_data(date, count, timezone='UTC+8'):
    '''
    生成一缓签到数据，供测试用，最多一百组数据
    注意，此函数不会返回指定数量的数据，
    :param date: 日期，可以是日期时间
    :param count: 数量
    :param timezone: 时区，不指定的话，默认UTC+8
    :return:
    '''
    import random

    user_ids = [
        33190536863, 46817092476, 56367541715, 32225839334, 29747380211, 74494408114, 77093027859, 76412956604,
        62623303615, 40280563259, 34739755100, 70892392576, 73186015041, 30332291113, 65250695812, 28023015340,
        56685741755, 75833358417, 44928644908, 71647804031, 45790793012, 41903239684, 30057838045, 31208252599,
        29202524460, 83896607851, 63373552895, 67863858729, 59573558877, 82077851178, 68921141536, 36109916355,
        84306372366, 55957208556, 85220080061, 57836864938, 64239520907, 43250388303, 57737211284, 51479134446,
        34644343464, 36714990541, 54361990918, 25757217282, 46094046811, 78533016098, 66567814863, 55213026411,
        30776864520, 52793241455, 62055690688, 40036868001, 69470917988, 44732102556, 73101954848, 67616510809,
        39813717540, 59827262161, 47393568359, 35835854225, 37415987291, 75371964621, 60137538183, 52508093850,
        38033613220, 67727511829, 52439264230, 63495446508, 33414814900, 79043236184, 54543913702, 29547115154,
        33985871398, 25479789589, 70273300936, 60833894362, 82525203384, 73760626352, 38539262703, 29695095787,
        64146659848, 71283264170, 69471295164, 62594689331, 43138813167, 26995161100, 32751043117, 48717510612,
        52758885935, 67894098797, 43468461924, 36970634128, 51513848035, 76194804578, 60375530688, 64189486459,
        51259679729, 82841899778, 42064616052, 62012878323
    ]
    name = ("天高任我飞,网约车,Abigail Ruiz,日入2000+ 支付宝项目,社三年级,蜡笔小新,派大星,HR-秋辞🔥信誉直招,唐老鸭一手cvv,特权,樂2.0,Mill,阿基米德,海绵宝宝,"
            "鼎富 二十一,Kelly Brown,萧墙之祸,刘尔丝,金革之难,平力浮,功不补患,端木琦杏,空音游,糟老头儿,杨敏达,新澳6合彩内幕12码,芳草,筱瑜,币圈奇哥,哼我是雄狮,"
            "可乐妹妹,帅帝,想发达,物联科技,品茶修车专线,锅内境外各种app接🐎成品主业对接有群,快活林 美女简介,sik jiang,阿巴阿巴 阿巴阿巴,青云,雾酱,马苏亚,雷吉吉,"
            "雨桐🍓清纯学妹,钵钵鸡,闷声 发大财,阿锐,刁民70681,门神,傑森,黎明明,钱顺,光圣,凯迪不拉客,驴中楚,卡比獸,鼎盛娱乐,长河月落,南门湾,壺齊𫵷,士大夫,青草仙,"
            "雾里看,雅芙（有课室,龙先生,雨泣,硬汉阿诺,夏总,阿迪耐克全球招代理,小钟3,龙冰语,多次拒绝卡戴珊,陈辰,Linh Phương,黑桃老K,Hải Nam,夜斗大弟壶,新手,"
            "王小美,查斯,,多多钱,一只小羊,,,天理,大大发,赛博终端,一路向钱,中午要睡觉,揽胜,婉晴,大雄引流,暖暖,专家打洞,多啦A梦,威龙,拜金女1 没登录,树形图设计者,"
            "luyang,发信,川普选妃,Makmadl,麻将胡了,杨厨,默默,雪梨")
    curr_unix = timestand.format_unix(date, timezone)
    name = name.split(',')
    result = {}
    for i, user in enumerate(user_ids[:count]):
        skip = random.randint(0, 10)
        if skip > 9:
            continue
        node_time = random.randint(curr_unix, curr_unix + 86400)
        result.update({
            user: {
                'register': [timestand.format_datetime_tz(node_time)],
                'first_name': name[i]
            }
        })

    return result


register = Register()

if __name__ == '__main__':



    pass
















